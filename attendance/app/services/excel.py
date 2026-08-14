"""회계사 전달용 엑셀(xlsx) 생성."""

from __future__ import annotations

import datetime as dt
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import STATUS_ANOMALY, STATUS_COMPLETE, STATUS_OPEN, Employee, WorkDay
from app.utils.timeutil import format_hhmm, format_time

_WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]

_STATUS_LABEL = {
    STATUS_COMPLETE: "정상",
    STATUS_OPEN: "퇴근 미기록",
    STATUS_ANOMALY: "이상",
}

_HEADER_FILL = PatternFill("solid", fgColor="1C1C1C")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WARN_FILL = PatternFill("solid", fgColor="FFF3CD")

_COLUMNS = [
    ("직원명", 14),
    ("사번", 10),
    ("날짜", 12),
    ("요일", 6),
    ("출근시각", 10),
    ("퇴근시각", 10),
    ("총 근무시간", 12),
    ("근무시간(소수)", 14),
    ("상태", 12),
    ("비고", 40),
]


def _style_header(sheet, columns: list[tuple[str, int]]) -> None:
    sheet.append([name for name, _ in columns])
    for index, (_, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"


def build_workbook(session: Session, start: dt.date, end: dt.date) -> bytes:
    """기간 내 근무기록 + 직원별 합계 두 시트를 담은 xlsx 바이트를 반환한다."""
    rows = session.execute(
        select(WorkDay, Employee)
        .join(Employee, Employee.id == WorkDay.employee_id)
        .where(WorkDay.work_date >= start, WorkDay.work_date <= end)
        .order_by(Employee.name, WorkDay.work_date)
    ).all()

    workbook = Workbook()
    detail = workbook.active
    detail.title = "근무기록"
    _style_header(detail, _COLUMNS)

    totals: dict[int, dict] = {}

    for work_day, employee in rows:
        minutes = work_day.worked_minutes
        detail.append(
            [
                employee.name,
                employee.code or "",
                work_day.work_date.strftime("%Y-%m-%d"),
                _WEEKDAYS[work_day.work_date.weekday()],
                format_time(work_day.check_in_at),
                format_time(work_day.check_out_at),
                format_hhmm(minutes),
                round(minutes / 60, 2),
                _STATUS_LABEL.get(work_day.status, work_day.status),
                work_day.note or "",
            ]
        )
        if work_day.needs_review or work_day.status != STATUS_COMPLETE:
            for column in range(1, len(_COLUMNS) + 1):
                detail.cell(row=detail.max_row, column=column).fill = _WARN_FILL

        bucket = totals.setdefault(
            employee.id,
            {"name": employee.name, "code": employee.code or "", "days": 0, "minutes": 0, "review": 0},
        )
        bucket["days"] += 1
        bucket["minutes"] += minutes
        if work_day.needs_review:
            bucket["review"] += 1

    summary_columns = [("직원명", 14), ("사번", 10), ("근무일수", 10), ("총 근무시간", 12), ("근무시간(소수)", 14), ("검토필요", 10)]
    summary = workbook.create_sheet("직원별 합계")
    _style_header(summary, summary_columns)
    for bucket in sorted(totals.values(), key=lambda b: b["name"]):
        summary.append(
            [
                bucket["name"],
                bucket["code"],
                bucket["days"],
                format_hhmm(bucket["minutes"]),
                round(bucket["minutes"] / 60, 2),
                bucket["review"],
            ]
        )

    summary.append([])
    summary.append([f"기간: {start} ~ {end}"])
    summary.append(["※ 시각은 사진 속 현장 보드 표시 시각 기준입니다."])
    summary.append(["※ 노란색 표시 행은 확인이 필요한 건입니다."])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def month_range(year: int, month: int) -> tuple[dt.date, dt.date]:
    start = dt.date(year, month, 1)
    end = dt.date(year + (month == 12), (month % 12) + 1, 1) - dt.timedelta(days=1)
    return start, end


def filename_for(start: dt.date, end: dt.date) -> str:
    if start.year == end.year and start.month == end.month and start.day == 1:
        return f"attendance_{start:%Y-%m}.xlsx"
    return f"attendance_{start:%Y%m%d}_{end:%Y%m%d}.xlsx"
